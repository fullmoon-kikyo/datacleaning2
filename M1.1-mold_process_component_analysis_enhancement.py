# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable=None, *args, **kwargs):
        return iterable


DEFAULT_591E_CANDIDATES = [
    Path("1-591E20260520-1.xlsx"),
    Path("1-591E20260414.xlsx"),
    Path("591E20260409.xlsx"),
]
DEFAULT_MATERIAL_LIST = Path("2-物料列表.xlsx")
DEFAULT_NITRIDING_DETAIL = Path("3-氮化明细.xlsx")
OUTPUT_FILE_PREFIX = "处理后模具数据"

MAIN_SHEET = "Data"
PROCESS_SUFFIXES = {"MC", "MR", "MN", "MT"}
MATERIAL_COLUMNS = ["物料编码", "物料号", "物料"]
FREEZE_STATUS_COLUMNS = ["基本视图状态"]
BOM_QTY_COLUMNS = ["BOM基本数量"]
MOLD_COLUMNS = ["模具号"]

PRODUCT_COL = "成品物料号"
NEW_MATERIAL_COL = "是否为新增物料"
ROUGH_COL = "粗加工"
HEAT_COL = "热处理"
NITRIDING_COL = "氮化"
CHILD_COL = "子件号"
PROCESS_COL = "工序"
S4_COL = "S4物料号"
FROZEN_COL = "是否冻结物料"
MISSING_BOM_COL = "是否缺少BOM信息"
NITRIDING_RECORD_COL = "是否存在氮化记录"

PROCESS_COLUMNS = {
    "MC": ROUGH_COL,
    "MR": HEAT_COL,
    "MN": NITRIDING_COL,
    "MT": NITRIDING_COL,
}

DF2_COLUMNS = [
    PRODUCT_COL,
    f"{PRODUCT_COL}-物料冻结状态",
    f"{PRODUCT_COL}-物料BOM状态",
    NEW_MATERIAL_COL,
    ROUGH_COL,
    f"{ROUGH_COL}-物料冻结状态",
    f"{ROUGH_COL}-物料BOM状态",
    HEAT_COL,
    f"{HEAT_COL}-物料冻结状态",
    f"{HEAT_COL}-物料BOM状态",
    NITRIDING_COL,
    f"{NITRIDING_COL}-物料冻结状态",
    f"{NITRIDING_COL}-物料BOM状态",
]

RESULT_GROUP_FILLS = {
    (PRODUCT_COL, f"{PRODUCT_COL}-物料冻结状态", f"{PRODUCT_COL}-物料BOM状态"): "DDEBF7",
    (ROUGH_COL, f"{ROUGH_COL}-物料冻结状态", f"{ROUGH_COL}-物料BOM状态"): "E2F0D9",
    (HEAT_COL, f"{HEAT_COL}-物料冻结状态", f"{HEAT_COL}-物料BOM状态"): "FFF2CC",
    (NITRIDING_COL, f"{NITRIDING_COL}-物料冻结状态", f"{NITRIDING_COL}-物料BOM状态"): "FCE4D6",
}
PROCESS_ROW_FILL_COLOR = "E7E6E6"


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def log(message: str) -> None:
    print(message, flush=True)


def progress(iterable, desc: str, total: int | None = None, colour: str = "green"):
    return tqdm(
        iterable,
        desc=desc,
        total=total,
        unit="行",
        dynamic_ncols=True,
        file=sys.stdout,
        colour=colour,
    )


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def find_column(df: pd.DataFrame, candidates: list[str], table_name: str, required: bool = True) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    if required:
        raise KeyError(f"{table_name} 缺少必要列，候选列名: {candidates}")
    return None


def is_status_09(value: object) -> bool:
    return clean_text(value) in {"09", "9", "9.0"}


def is_zero(value: object) -> bool:
    if pd.isna(value):
        return False
    try:
        return float(value) == 0
    except (TypeError, ValueError):
        return clean_text(value) == "0"


def split_material_code(value: object) -> tuple[str, str]:
    text = clean_text(value)
    suffix = text[-2:].upper() if len(text) >= 2 else ""
    if suffix in PROCESS_SUFFIXES:
        return text[:-2], suffix
    return text, ""


def build_s4_material(value: object) -> str:
    mold_no = clean_text(value)
    return f"S-{mold_no}" if mold_no else ""


def add_blank_columns(df: pd.DataFrame, columns: list[str]) -> None:
    for col in columns:
        df[col] = ""


def first_existing_path(candidates: list[Path]) -> Path:
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


def build_output_path(output: str | None = None) -> Path:
    if output:
        return Path(output)
    return Path(f"{OUTPUT_FILE_PREFIX}_{datetime.now():%Y%m%d}.xlsx")


def read_excel_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    excel = pd.ExcelFile(path)
    selected_sheet = sheet_name if sheet_name in excel.sheet_names else 0
    if selected_sheet == 0 and sheet_name not in excel.sheet_names:
        log(f"{path.name} 未找到工作表 {sheet_name}，改为读取第一个工作表。")
    return pd.read_excel(excel, sheet_name=selected_sheet)


def create_row_index(df: pd.DataFrame, key_col: str, desc: str) -> dict[str, int]:
    index_by_key: dict[str, int] = {}
    for row_idx in progress(df.index, desc=desc, total=len(df), colour="blue"):
        key = clean_text(df.at[row_idx, key_col])
        if key and key not in index_by_key:
            index_by_key[key] = row_idx
    return index_by_key


def mark_df1_status(
    df1: pd.DataFrame,
    table_name: str,
    freeze_col: str | None,
    bom_qty_col: str | None,
) -> None:
    add_blank_columns(df1, [FROZEN_COL, MISSING_BOM_COL])

    if freeze_col is None:
        log(f"{table_name} 未找到基本视图状态列，跳过冻结状态标记。")
    if bom_qty_col is None:
        log(f"{table_name} 未找到 BOM 基本数量列，跳过缺少 BOM 信息标记。")

    if freeze_col is None and bom_qty_col is None:
        return

    for row_idx in progress(df1.index, desc="标记冻结/BOM状态", total=len(df1), colour="cyan"):
        if freeze_col is not None and is_status_09(df1.at[row_idx, freeze_col]):
            df1.at[row_idx, FROZEN_COL] = "F"
        if bom_qty_col is not None and is_zero(df1.at[row_idx, bom_qty_col]):
            df1.at[row_idx, MISSING_BOM_COL] = "F"


def split_df1_material_codes(df1: pd.DataFrame, material_col: str) -> None:
    add_blank_columns(df1, [CHILD_COL, PROCESS_COL])

    for row_idx in progress(df1.index, desc="拆分物料编码", total=len(df1), colour="cyan"):
        child_no, process = split_material_code(df1.at[row_idx, material_col])
        df1.at[row_idx, CHILD_COL] = child_no
        df1.at[row_idx, PROCESS_COL] = process


def append_new_materials(df2: pd.DataFrame, material_records: pd.DataFrame) -> pd.DataFrame:
    existing = set()
    for row_idx in progress(df2.index, desc="建立处理结果子件号索引", total=len(df2), colour="blue"):
        child_no = clean_text(df2.at[row_idx, CHILD_COL])
        if child_no:
            existing.add(child_no)

    rows_to_append: list[dict[str, str]] = []
    for row_idx in progress(material_records.index, desc="比对新增物料", total=len(material_records), colour="yellow"):
        s4_material = clean_text(material_records.at[row_idx, S4_COL])
        if s4_material and s4_material not in existing:
            new_row = {col: "" for col in df2.columns}
            new_row[CHILD_COL] = s4_material
            new_row[NEW_MATERIAL_COL] = "是"
            rows_to_append.append(new_row)
            existing.add(s4_material)

    if rows_to_append:
        log(f"发现新增物料 {len(rows_to_append)} 条，追加到处理结果。")
        df2 = pd.concat([df2, pd.DataFrame(rows_to_append)], ignore_index=True)
    else:
        log("未发现需要追加的新增物料。")
    return df2


def fill_process_columns(df1: pd.DataFrame, df2: pd.DataFrame, material_col: str) -> None:
    row_index_by_child = create_row_index(df2, CHILD_COL, "建立处理结果子件号索引")

    for row_idx in progress(df1.index, desc="回填工序与状态", total=len(df1), colour="green"):
        child_no, process = split_material_code(df1.at[row_idx, material_col])
        if not child_no or child_no not in row_index_by_child:
            continue

        target_idx = row_index_by_child[child_no]
        frozen_status = df1.at[row_idx, FROZEN_COL]
        bom_status = df1.at[row_idx, MISSING_BOM_COL]

        if process in PROCESS_COLUMNS:
            target_col = PROCESS_COLUMNS[process]
            df2.at[target_idx, target_col] = process
            df2.at[target_idx, f"{target_col}-物料冻结状态"] = frozen_status
            df2.at[target_idx, f"{target_col}-物料BOM状态"] = bom_status
        elif process == "":
            df2.at[target_idx, PRODUCT_COL] = child_no
            df2.at[target_idx, f"{PRODUCT_COL}-物料冻结状态"] = frozen_status
            df2.at[target_idx, f"{PRODUCT_COL}-物料BOM状态"] = bom_status


def add_s4_material_column(df: pd.DataFrame, table_name: str) -> None:
    mold_col = find_column(df, MOLD_COLUMNS, table_name)
    df[S4_COL] = ""
    for row_idx in progress(df.index, desc=f"生成{table_name} S4物料号", total=len(df), colour="magenta"):
        df.at[row_idx, S4_COL] = build_s4_material(df.at[row_idx, mold_col])


def fill_nitriding_record(df2: pd.DataFrame, nitriding_records: pd.DataFrame) -> None:
    df2[NITRIDING_RECORD_COL] = ""

    nitriding_index_by_s4 = create_row_index(nitriding_records, S4_COL, "建立氮化记录索引")
    matched_count = 0
    for row_idx in progress(df2.index, desc="匹配氮化记录", total=len(df2), colour="yellow"):
        child_no = clean_text(df2.at[row_idx, CHILD_COL])
        if child_no and child_no in nitriding_index_by_s4:
            source_idx = nitriding_index_by_s4[child_no]
            df2.at[row_idx, NITRIDING_RECORD_COL] = nitriding_records.at[source_idx, S4_COL]
            matched_count += 1
    log(f"氮化记录匹配完成，共匹配 {matched_count} 条。")


def beautify_result_sheet(ws) -> None:
    log("正在美化处理结果: 设置分组表头颜色。")
    headers = {cell.value: cell.column for cell in ws[1]}

    for columns, fill_color in RESULT_GROUP_FILLS.items():
        fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
        for column_name in columns:
            col_idx = headers.get(column_name)
            if col_idx is not None:
                ws.cell(row=1, column=col_idx).fill = fill

    log("正在美化处理结果: 添加工序非空整行染色条件格式。")
    if ws.max_row < 2:
        log("处理结果无数据行，跳过整行染色规则。")
        return

    process_cols = [ROUGH_COL, HEAT_COL, NITRIDING_COL]
    missing_cols = [column_name for column_name in process_cols if column_name not in headers]
    if missing_cols:
        log(f"处理结果缺少列 {missing_cols}，跳过整行染色规则。")
        return

    row_fill = PatternFill(fill_type="solid", start_color=PROCESS_ROW_FILL_COLOR, end_color=PROCESS_ROW_FILL_COLOR)
    rough_letter = get_column_letter(headers[ROUGH_COL])
    heat_letter = get_column_letter(headers[HEAT_COL])
    nitriding_letter = get_column_letter(headers[NITRIDING_COL])
    data_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    formula = f'OR(${rough_letter}2<>"",${heat_letter}2<>"",${nitriding_letter}2<>"")'
    ws.conditional_formatting.add(data_range, FormulaRule(formula=[formula], fill=row_fill))
    log("处理结果整行染色规则已添加。")


def write_excel(
    output_path: Path,
    df1: pd.DataFrame,
    material_records: pd.DataFrame,
    df2: pd.DataFrame,
    nitriding_records: pd.DataFrame,
    apply_format: bool = True,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        log("正在写入 sheet: Data")
        df1.to_excel(writer, sheet_name="Data", index=False)
        log("正在写入 sheet: 物料生产记录")
        material_records.to_excel(writer, sheet_name="物料生产记录", index=False)
        log("正在写入 sheet: 处理结果")
        df2.to_excel(writer, sheet_name="处理结果", index=False)
        log("正在写入 sheet: 氮化记录")
        nitriding_records.to_excel(writer, sheet_name="氮化记录", index=False)
        if apply_format:
            beautify_result_sheet(writer.sheets["处理结果"])


def write_output(
    output_path: Path,
    df1: pd.DataFrame,
    material_records: pd.DataFrame,
    df2: pd.DataFrame,
    nitriding_records: pd.DataFrame,
    apply_format: bool,
) -> Path:
    actual_output = output_path
    try:
        write_excel(actual_output, df1, material_records, df2, nitriding_records, apply_format)
    except PermissionError:
        actual_output = actual_output.with_name(
            f"{actual_output.stem}_{datetime.now():%Y%m%d_%H%M%S}{actual_output.suffix}"
        )
        log(f"无法覆盖输出文件，文件可能正在打开，改为另存为: {actual_output}")
        write_excel(actual_output, df1, material_records, df2, nitriding_records, apply_format)
    return actual_output


def parse_args() -> argparse.Namespace:
    default_input = first_existing_path(DEFAULT_591E_CANDIDATES)
    parser = argparse.ArgumentParser(description="清洗模具物料数据并生成过程组件分析结果")
    parser.add_argument("--input", default=str(default_input), help=f"591E 输入文件，默认: {default_input}")
    parser.add_argument("--sheet", default=MAIN_SHEET, help=f"591E 工作表名，默认: {MAIN_SHEET}")
    parser.add_argument("--material-list", default=str(DEFAULT_MATERIAL_LIST), help="物料列表文件")
    parser.add_argument("--nitriding-detail", default=str(DEFAULT_NITRIDING_DETAIL), help="氮化明细文件")
    parser.add_argument("--output", default=None, help="输出文件名，默认按日期生成")
    parser.add_argument("--no-format", action="store_true", help="不对输出 Excel 添加颜色和条件格式")
    return parser.parse_args()


def main() -> None:
    configure_console_encoding()
    args = parse_args()

    input_591e = Path(args.input)
    material_list_path = Path(args.material_list)
    nitriding_detail_path = Path(args.nitriding_detail)
    output_path = build_output_path(args.output)

    log("开始读取 Excel 文件。")
    log(f"591E 输入文件: {input_591e.resolve()}")
    log(f"物料列表文件: {material_list_path.resolve()}")
    log(f"氮化明细文件: {nitriding_detail_path.resolve()}")

    df1 = read_excel_sheet(input_591e, args.sheet)
    material_records = pd.read_excel(material_list_path, dtype={"模具号": str})
    nitriding_records = pd.read_excel(nitriding_detail_path, dtype={"模具号": str})

    material_col = find_column(df1, MATERIAL_COLUMNS, input_591e.name)
    freeze_col = find_column(df1, FREEZE_STATUS_COLUMNS, input_591e.name, required=False)
    bom_qty_col = find_column(df1, BOM_QTY_COLUMNS, input_591e.name, required=False)
    find_column(material_records, MOLD_COLUMNS, material_list_path.name)
    find_column(nitriding_records, MOLD_COLUMNS, nitriding_detail_path.name)

    log("处理 Data: 标记冻结/BOM状态，并拆分物料编码。")
    mark_df1_status(df1, input_591e.name, freeze_col, bom_qty_col)
    split_df1_material_codes(df1, material_col)

    log("生成处理结果: 提取子件号去重并增加结果列。")
    df2 = df1[[CHILD_COL]].drop_duplicates().reset_index(drop=True)
    add_blank_columns(df2, DF2_COLUMNS)

    log("处理物料列表: 生成 S4物料号并追加新增物料。")
    add_s4_material_column(material_records, material_list_path.name)
    df2 = append_new_materials(df2, material_records)

    log("回填工序、冻结状态和 BOM 状态。")
    fill_process_columns(df1, df2, material_col)

    log("处理氮化明细: 生成 S4物料号并匹配氮化记录。")
    add_s4_material_column(nitriding_records, nitriding_detail_path.name)
    fill_nitriding_record(df2, nitriding_records)

    log("写入输出文件。")
    actual_output = write_output(
        output_path,
        df1,
        material_records,
        df2,
        nitriding_records,
        apply_format=not args.no_format,
    )

    log("处理完成。")
    log(f"输出文件: {actual_output.resolve()}")
    log(f"Data 行数: {len(df1)}")
    log(f"物料生产记录 行数: {len(material_records)}")
    log(f"处理结果 行数: {len(df2)}")
    log(f"氮化记录 行数: {len(nitriding_records)}")


if __name__ == "__main__":
    main()
