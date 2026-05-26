# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm


BASE_FIELDS = [
    "描述",
    "物料类型",
    "物料组",
    "库位",
    "主数据数量",
    "主数据单位",
    "BOM单位",
    "BOM数量",
    "工艺路线数量",
    "分子",
    "分母",
]

INPUT_SHEET = "Data"
INPUT_FILE = "591E20260409-有BOM"
OUTPUT_FILE = "1511-转换关系整合.xlsx"
NUMERIC_COL_KEYWORDS = ("数量", "分子", "分母")


def beautify_sheet(ws) -> None:
    # 基础视图
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 110

    # 表头样式
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = {}
    for cell in ws[1]:
        headers[cell.value] = cell.column
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 列宽调整（基于前N行采样，避免超大文件过慢）
    sample_max_row = min(ws.max_row, 3000)
    for col_cells in tqdm(
        ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=sample_max_row),
        total=ws.max_column,
        desc="美化-列宽",
    ):
        col_idx = col_cells[0].column
        col_name = col_cells[0].value
        col_letter = get_column_letter(col_idx)

        if col_name in {"物料"}:
            ws.column_dimensions[col_letter].width = 18
            continue
        if col_name in {"1-描述", "2-描述", "描述列"}:
            ws.column_dimensions[col_letter].width = 28
            continue

        max_len = len(str(col_name)) if col_name is not None else 8
        for cell in col_cells[1:]:
            value = cell.value
            if value is None:
                continue
            v_len = len(str(value))
            if v_len > max_len:
                max_len = v_len

        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 24)

    # 列对齐 + 数字格式
    left_cols = {"物料", "1-描述", "2-描述", "描述列"}
    center_cols = {"是否异常", "转换关系异常"}
    # 大表时避免对所有列逐单元格设样式，优先保证关键列和数值列。
    full_alignment = ws.max_row <= 30000
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    for col_idx in tqdm(range(1, ws.max_column + 1), total=ws.max_column, desc="美化-格式"):
        col_name = ws.cell(row=1, column=col_idx).value
        is_numeric = any(k in str(col_name) for k in NUMERIC_COL_KEYWORDS)
        is_left = col_name in left_cols
        is_center = col_name in center_cols
        if is_left:
            target_align = left_align
        elif is_center or full_alignment:
            target_align = center_align
        else:
            target_align = None

        if not is_numeric and target_align is None:
            continue

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if is_numeric and cell.value is not None:
                cell.number_format = "General"
            if target_align is not None:
                cell.alignment = target_align


def find_input_file(input_name: str) -> Path:
    p = Path(input_name)
    if p.exists():
        return p

    if p.suffix == "":
        for ext in (".xlsx", ".xlsm", ".xls"):
            candidate = Path(f"{input_name}{ext}")
            if candidate.exists():
                return candidate

    direct_matches = sorted(Path(".").glob(f"{input_name}*.xls*"))
    if direct_matches:
        return direct_matches[0]

    # 兜底: 用户常会写成 "xxxx-有BOM"，但实际文件名可能只有前缀。
    prefix = input_name.split("-")[0]
    fallback_matches = sorted(Path(".").glob(f"{prefix}*.xls*"))
    if fallback_matches:
        return fallback_matches[0]

    raise FileNotFoundError(f"未找到输入文件: {input_name}")


def compute_conversion_abnormal_mask(df: pd.DataFrame) -> pd.Series:
    keys = ["1-BOM数量", "1-工艺路线数量", "2-主数据数量", "2-BOM数量", "2-工艺路线数量"]
    vals = df[keys]
    complete = vals.notna().all(axis=1)

    result = pd.Series(False, index=df.index, dtype=bool)
    if not complete.any():
        return result

    complete_vals = vals.loc[complete]
    numeric = complete_vals.apply(pd.to_numeric, errors="coerce")
    numeric_rows = numeric.notna().all(axis=1)

    if numeric_rows.any():
        numeric_vals = numeric.loc[numeric_rows]
        diff = numeric_vals.sub(numeric_vals[keys[0]], axis=0).abs()
        numeric_abnormal = diff.iloc[:, 1:].gt(1e-9).any(axis=1)
        result.loc[numeric_abnormal.index] = numeric_abnormal

    non_numeric_idx = complete_vals.index[~numeric_rows]
    if len(non_numeric_idx) > 0:
        str_vals = complete_vals.loc[non_numeric_idx].astype(str).apply(lambda s: s.str.strip())
        string_abnormal = str_vals.iloc[:, 1:].ne(str_vals[keys[0]], axis=0).any(axis=1)
        result.loc[string_abnormal.index] = string_abnormal

    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="整合EA/ZM转换关系并标记异常")
    parser.add_argument(
        "--input",
        default=INPUT_FILE,
        help=f"输入文件名(可不带扩展名)，默认: {INPUT_FILE}",
    )
    parser.add_argument(
        "--sheet",
        default=INPUT_SHEET,
        help=f"工作表名称，默认: {INPUT_SHEET}",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT_FILE,
        help=f"输出文件名，默认: {OUTPUT_FILE}",
    )
    args = parser.parse_args()

    input_path = find_input_file(args.input)
    output_path = Path(args.output)

    print(f"读取文件: {input_path}")
    df1 = pd.read_excel(input_path, sheet_name=args.sheet)

    required_cols = ["物料", "主数据单位", *BASE_FIELDS]
    missing = [c for c in required_cols if c not in df1.columns]
    if missing:
        raise KeyError(f"Data页缺少必要列: {missing}")

    print("步骤1/5: 提取唯一物料")
    df2 = pd.DataFrame({"物料": df1["物料"].drop_duplicates().reset_index(drop=True)})

    print("步骤2/5: 构建EA/ZM映射并整合")
    unit_df = df1[["物料", *BASE_FIELDS]].copy()
    unit_df["主数据单位"] = unit_df["主数据单位"].astype(str).str.strip().str.upper()
    unit_df = unit_df[unit_df["主数据单位"].isin({"EA", "ZM"})]
    unit_df = unit_df.drop_duplicates(subset=["物料", "主数据单位"], keep="first")

    ea_df = (
        unit_df[unit_df["主数据单位"] == "EA"][["物料", *BASE_FIELDS]]
        .rename(columns={field: f"1-{field}" for field in BASE_FIELDS})
    )
    zm_df = (
        unit_df[unit_df["主数据单位"] == "ZM"][["物料", *BASE_FIELDS]]
        .rename(columns={field: f"2-{field}" for field in BASE_FIELDS})
    )
    df2 = df2.merge(ea_df, on="物料", how="left").merge(zm_df, on="物料", how="left")
    df2["是否异常"] = pd.NA
    df2["描述列"] = pd.NA
    df2["转换关系异常"] = pd.NA

    print("步骤3/5: 判定异常")
    has_ea = df2["1-主数据单位"].notna()
    has_zm = df2["2-主数据单位"].notna()

    missing_pair = ~(has_ea & has_zm)
    df2.loc[missing_pair, "是否异常"] = 1
    df2.loc[~has_ea & has_zm, "描述列"] = "缺失EA"
    df2.loc[has_ea & ~has_zm, "描述列"] = "缺失ZM"
    df2.loc[~has_ea & ~has_zm, "描述列"] = "缺失EA和ZM"

    conversion_abnormal = compute_conversion_abnormal_mask(df2)
    df2.loc[conversion_abnormal, "转换关系异常"] = 1

    print("步骤4/5: 导出并美化")

    def write_workbook(target_path: Path) -> None:
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            df2.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]

            print("步骤5/5: 美化并标色")
            beautify_sheet(ws)

            yellow_fill = PatternFill(fill_type="solid", start_color="FFF59D", end_color="FFF59D")
            red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

            headers = {cell.value: cell.column for cell in ws[1]}
            abnormal_col = headers["是否异常"]
            conv_col = headers["转换关系异常"]
            if ws.max_row >= 2:
                data_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
                abnormal_letter = get_column_letter(abnormal_col)
                conv_letter = get_column_letter(conv_col)

                red_rule = FormulaRule(
                    formula=[f'OR(${conv_letter}2=1,${conv_letter}2="1",${conv_letter}2="1.0")'],
                    fill=red_fill,
                    stopIfTrue=True,
                )
                yellow_rule = FormulaRule(
                    formula=[f'OR(${abnormal_letter}2=1,${abnormal_letter}2="1",${abnormal_letter}2="1.0")'],
                    fill=yellow_fill,
                )

                ws.conditional_formatting.add(data_range, red_rule)
                ws.conditional_formatting.add(data_range, yellow_rule)

    try:
        write_workbook(output_path)
        final_output = output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        final_output = output_path.with_name(f"{output_path.stem}-{timestamp}{output_path.suffix}")
        write_workbook(final_output)
        print(f"警告: 文件被占用，已另存为 {final_output}")

    print(f"完成: {final_output.resolve()}")


if __name__ == "__main__":
    main()
