# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter


STOCK_FILE_PATTERN = "01_3005*.xlsx"
STOCK_SHEET_NAME = "Data"
BASE_DIR = Path(__file__).resolve().parent
COUNT_DATA_DIR = BASE_DIR / "02_盘点数据"
COUNT_SHEET_NAME = "Sheet1"

STOCK_COLUMN_ALIASES = [
    ("物料", ["物料"]),
    ("物料描述", ["物料描述"]),
    ("批次", ["批次"]),
    ("库存地点", ["库存地点", "存储地点"]),
    ("非限制使用的库存", ["非限制使用的库存"]),
    ("单位", ["单位", "基本计量单位"]),
    ("数量(生产单位)", ["数量(生产单位)", "数量（生产单位）"]),
    ("生产单位", ["生产单位"]),
]

DF5_GROUP_COLORS = [
    "9CC2E5",
    "F4B183",
    "A9D18E",
    "B4A7D6",
    "F4B6C2",
    "76D7C4",
]

DF5_SUMMARY_COLORS = [
    "BDD7EE",
    "F8CBAD",
    "C6E0B4",
    "D9D2E9",
    "F4CCCC",
    "B7E1D8",
]

DF5_SYNC_MATCH_COLORS = [
    "00FFFF",
    "00FF00",
    "FF00FF",
    "00B0F0",
    "92D050",
    "FF66CC",
    "00FF99",
    "7030A0",
    "FF9900",
    "33CCCC",
    "CC00FF",
    "66FF33",
    "FF5050",
    "0099FF",
    "CCFF00",
    "FF33CC",
]
DF5_SYNC_STOCK_ONLY_COLOR = "FFFF00"
DF5_SYNC_PHYSICAL_ONLY_COLOR = "FFFF00"


def configure_console_encoding() -> None:
    """尽量保证 Windows 控制台中文输出正常。"""
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def log(message: str) -> None:
    print(message, flush=True)


def clean_text(value: object) -> str:
    """将 Excel 单元格内容转为干净文本，并处理 123.0 这类整数浮点显示。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_simple_manufacture_number(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if text.isdigit():
        return text
    if re.fullmatch(r"[A-Za-z0-9]+", text) and any(char.isdigit() for char in text) and any(
        char.isalpha() for char in text
    ):
        six_digit_match = re.search(r"\d{6}", text)
        if six_digit_match:
            return six_digit_match.group(0)
    return text


def find_latest_stock_file() -> Path:
    candidates = [
        path
        for path in BASE_DIR.glob(STOCK_FILE_PATTERN)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"未找到库存文件：{STOCK_FILE_PATTERN}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def resolve_count_data_dir() -> Path:
    if COUNT_DATA_DIR.exists():
        return COUNT_DATA_DIR

    matching_dirs = [
        path
        for path in BASE_DIR.glob("*")
        if path.is_dir() and "盘点" in path.name and "数据" in path.name
    ]
    if matching_dirs:
        selected_dir = max(matching_dirs, key=lambda path: path.stat().st_mtime)
        log(f"未找到默认盘点数据文件夹：{COUNT_DATA_DIR.name}，改用：{selected_dir.name}")
        return selected_dir

    raise FileNotFoundError(f"未找到盘点数据文件夹：{COUNT_DATA_DIR}")


def list_count_files() -> list[Path]:
    count_data_dir = resolve_count_data_dir()

    files = sorted(
        path
        for path in count_data_dir.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(f"{count_data_dir} 下未找到 .xlsx 文件")
    return files


def extract_count_code(path: Path) -> str:
    """
    从文件名提取盘点标识。
    例如：02_实物盘点-NY-20260617.xlsx -> NY
    """
    parts = path.stem.split("-")
    if len(parts) >= 3:
        return parts[1].strip() or path.stem
    if len(parts) == 2:
        return parts[1].strip() or path.stem
    return path.stem


def validate_columns(df: pd.DataFrame, required_columns: list[str], table_name: str) -> None:
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise KeyError(f"{table_name} 缺少必要列：{missing}")


def build_material_code_from_drawing(value: object) -> str:
    drawing_number = re.sub(r"[A-Za-z]+$", "", clean_text(value))
    if not drawing_number:
        return ""
    if drawing_number.upper().startswith("D"):
        return f"S-{drawing_number}"
    return f"S-D{drawing_number}"


def build_df2(df1: pd.DataFrame) -> pd.DataFrame:
    missing: list[str] = []
    selected_columns: dict[str, pd.Series] = {}

    for output_column, candidates in STOCK_COLUMN_ALIASES:
        source_column = next((column for column in candidates if column in df1.columns), None)
        if source_column is None:
            missing.append(output_column)
        else:
            selected_columns[output_column] = df1[source_column]

    if missing:
        raise KeyError(f"库存 Data 页缺少必要列：{missing}")

    return pd.DataFrame(selected_columns)


def process_df3(df: pd.DataFrame, table_name: str) -> pd.DataFrame:
    validate_columns(df, ["图号", "制造号"], table_name)
    result = df.copy()

    material_codes = result["图号"].map(build_material_code_from_drawing)
    if "物料编码" in result.columns:
        result["物料编码"] = material_codes
    else:
        insert_at = result.columns.get_loc("图号") + 1
        result.insert(insert_at, "物料编码", material_codes)

    split_values = result["制造号"].map(lambda value: clean_text(value).split())
    max_split_count = int(split_values.map(len).max()) if not split_values.empty else 0

    for index in range(max_split_count):
        column_name = f"制造号-{index + 1}"
        result[column_name] = split_values.map(
            lambda values, position=index: values[position] if position < len(values) else ""
        )

    return result


def get_manufacture_columns(df: pd.DataFrame) -> list[str]:
    manufacture_columns = [
        column
        for column in df.columns
        if isinstance(column, str) and re.fullmatch(r"制造号-\d+", column)
    ]
    return sorted(manufacture_columns, key=lambda column: int(column.split("-")[-1]))


def build_df4(df3: pd.DataFrame, table_name: str) -> pd.DataFrame:
    validate_columns(df3, ["图号", "物料编码", "实物数量"], table_name)
    manufacture_columns = get_manufacture_columns(df3)

    rows: list[dict[str, object]] = []
    for _, row in df3.iterrows():
        drawing_number = clean_text(row["图号"])
        material_code = clean_text(row["物料编码"])
        if not material_code:
            continue

        actual_quantity = row["实物数量"]
        manufacture_numbers = [
            clean_text(row[column])
            for column in manufacture_columns
            if clean_text(row[column])
        ]

        if manufacture_numbers:
            rows.extend(
                {
                    "图号": drawing_number,
                    "物料编码": material_code,
                    "实物数量": 1,
                    "实物制造号": manufacture_number,
                }
                for manufacture_number in manufacture_numbers
            )
        else:
            rows.append(
                {
                    "图号": drawing_number,
                    "物料编码": material_code,
                    "实物数量": actual_quantity,
                    "实物制造号": "",
                }
            )

    return pd.DataFrame(rows, columns=["图号", "物料编码", "实物数量", "实物制造号"])


def build_df4_summary(df4_dataframes: dict[str, pd.DataFrame]) -> pd.DataFrame:
    columns = ["图号", "物料编码", "实物数量", "实物制造号"]
    if not df4_dataframes:
        return pd.DataFrame(columns=columns)
    return pd.concat(
        [df4.reindex(columns=columns) for df4 in df4_dataframes.values()],
        ignore_index=True,
    )


def make_blank_df5_row(columns: list[str]) -> dict[str, object]:
    return {column: "" for column in columns}


def fill_physical_fields(
    row: dict[str, object],
    drawing_number: str,
    material_code: str,
    quantity: object,
    manufacture_number: str,
    status: str,
) -> None:
    row["图号"] = drawing_number
    row["实物物料编码"] = material_code
    row["实物数量"] = quantity
    row["实物制造号"] = manufacture_number
    row["简易制造号2"] = build_simple_manufacture_number(manufacture_number)
    row["状态"] = status


def find_first_empty_status_index(rows: list[dict[str, object]], indexes: list[int]) -> int:
    for index in indexes:
        if not clean_text(rows[index].get("状态", "")):
            return index
    return indexes[0]


def get_df5_row_material_key(row: dict[str, object]) -> str:
    return clean_text(row.get("物料", "")) or clean_text(row.get("实物物料编码", ""))


def apply_status_to_material_groups(df5: pd.DataFrame) -> pd.DataFrame:
    result = df5.copy()
    stock_material_keys = result["物料"].map(clean_text)
    physical_material_keys = result["实物物料编码"].map(clean_text)
    material_keys = stock_material_keys.mask(stock_material_keys == "", physical_material_keys)

    grouped_indexes = result.index.to_series().groupby(material_keys, sort=False)
    for material_key, indexes in grouped_indexes:
        if material_key == "":
            continue
        group_rows = result.loc[indexes].to_dict("records")
        result.loc[indexes, "状态"] = determine_df5_group_status(group_rows)

    return result


def to_number(value: object) -> float:
    text = clean_text(value).replace(",", "")
    if not text:
        return 0.0
    number = pd.to_numeric(text, errors="coerce")
    return 0.0 if pd.isna(number) else float(number)


def is_nonzero_number(value: object) -> bool:
    text = clean_text(value).replace(",", "")
    if not text:
        return False
    number = pd.to_numeric(text, errors="coerce")
    return not pd.isna(number) and float(number) != 0


def clean_summary_number(value: float) -> float | int:
    return int(value) if float(value).is_integer() else value


def sum_group_column(rows: list[dict[str, object]], column: str) -> float:
    return sum(to_number(row.get(column, "")) for row in rows)


def convert_column_to_number(df: pd.DataFrame, column: str) -> None:
    if column in df.columns:
        df[column] = pd.to_numeric(
            df[column].map(lambda value: clean_text(value).replace(",", "")),
            errors="coerce",
        ).fillna(0)


def get_stock_rows(group_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return [row for row in group_rows if clean_text(row.get("数据类型", "")) == "账面"]


def get_count_rows(group_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return [row for row in group_rows if clean_text(row.get("数据类型", "")) == "盘点"]


def calculate_stock_summary_quantity(group_rows: list[dict[str, object]]) -> float:
    stock_rows = get_stock_rows(group_rows)
    stock_quantity = sum_group_column(stock_rows, "非限制使用的库存")
    if stock_quantity < 500:
        return stock_quantity

    if any(not is_nonzero_number(row.get("数量(生产单位)", "")) for row in stock_rows):
        return float(len(stock_rows))

    production_quantity = sum_group_column(stock_rows, "数量(生产单位)")
    return production_quantity


def get_group_quantity_summary(group_rows: list[dict[str, object]]) -> tuple[float, float, str]:
    stock_summary_quantity = calculate_stock_summary_quantity(group_rows)
    physical_quantity = sum_group_column(group_rows, "实物数量")

    if stock_summary_quantity > physical_quantity:
        consistency = "账面多"
    elif physical_quantity > stock_summary_quantity:
        consistency = "实物多"
    else:
        consistency = "数量一致"
    return stock_summary_quantity, physical_quantity, consistency


def are_all_stock_simple_numbers_in_count_rows(group_rows: list[dict[str, object]]) -> bool:
    stock_simple_number_values = [
        clean_text(row.get("简易制造号1", "")) for row in get_stock_rows(group_rows)
    ]
    if not stock_simple_number_values or any(not value for value in stock_simple_number_values):
        return False

    stock_simple_numbers = Counter(stock_simple_number_values)
    count_simple_numbers = Counter(
        value
        for row in get_count_rows(group_rows)
        if (value := clean_text(row.get("简易制造号2", "")))
    )

    for simple_number, stock_count in stock_simple_numbers.items():
        if count_simple_numbers[simple_number] < stock_count:
            return False
    return bool(stock_simple_numbers)


def determine_df5_group_status(group_rows: list[dict[str, object]]) -> str:
    stock_rows = get_stock_rows(group_rows)
    count_rows = get_count_rows(group_rows)

    if not stock_rows and count_rows:
        return "仅实物"
    if stock_rows and not count_rows:
        return "仅账面"
    if not stock_rows and not count_rows:
        return "仅账面"

    _, _, consistency = get_group_quantity_summary(group_rows)
    if consistency == "数量一致" and are_all_stock_simple_numbers_in_count_rows(group_rows):
        return "账实一致"
    return "批次号不匹配"


def make_df5_summary_row(columns: list[str], group_rows: list[dict[str, object]]) -> dict[str, object]:
    summary_row = make_blank_df5_row(columns)
    summary_row["数据类型"] = "汇总"
    summary_row["S4物料编码"] = next(
        (material_key for row in group_rows if (material_key := get_df5_row_material_key(row))),
        "",
    )
    summary_row["状态"] = determine_df5_group_status(group_rows)
    stock_summary_quantity, physical_quantity, consistency = get_group_quantity_summary(group_rows)

    summary_row["非限制使用的库存"] = clean_summary_number(stock_summary_quantity)
    summary_row["实物数量"] = clean_summary_number(physical_quantity)
    summary_row["数量一致性"] = consistency
    return summary_row


def apply_quantity_consistency_to_group(group_rows: list[dict[str, object]]) -> None:
    _, _, consistency = get_group_quantity_summary(group_rows)
    for row in group_rows:
        row["数量一致性"] = consistency


def get_manufacture_sort_value(value: object) -> tuple[int, int, int, str]:
    text = clean_text(value)
    if not text:
        return (1, 0, 0, "")
    if text.isdigit():
        return (0, 0, int(text), text)
    return (0, 1, 0, text)


def sort_df5_group_rows(group_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    data_type_order = {"账面": 0, "盘点": 1}

    def sort_key(indexed_row: tuple[int, dict[str, object]]) -> tuple[object, ...]:
        index, row = indexed_row
        data_type = clean_text(row.get("数据类型", ""))
        manufacture_column = "简易制造号1" if data_type == "账面" else "简易制造号2"
        return (
            data_type_order.get(data_type, 2),
            get_manufacture_sort_value(row.get(manufacture_column, "")),
            index,
        )

    return [row for _, row in sorted(enumerate(group_rows), key=sort_key)]


def apply_sync_numbers_to_group(group_rows: list[dict[str, object]]) -> None:
    stock_rows_by_simple_number: dict[str, list[dict[str, object]]] = {}

    for row in group_rows:
        row["同步1"] = ""
        row["同步2"] = ""
        if clean_text(row.get("数据类型", "")) != "账面":
            continue

        simple_number = clean_text(row.get("简易制造号1", ""))
        if simple_number:
            stock_rows_by_simple_number.setdefault(simple_number, []).append(row)

    sync_number = 1
    for count_row in group_rows:
        if clean_text(count_row.get("数据类型", "")) not in {"账面", "盘点"}:
            continue

        simple_number = clean_text(count_row.get("简易制造号2", ""))
        if not simple_number:
            continue

        stock_rows = stock_rows_by_simple_number.get(simple_number, [])
        if stock_rows:
            own_stock_index = next(
                (
                    index
                    for index, stock_row in enumerate(stock_rows)
                    if stock_row is count_row
                ),
                None,
            )
            stock_row = stock_rows.pop(own_stock_index if own_stock_index is not None else 0)
            stock_row["同步1"] = sync_number
            count_row["同步2"] = sync_number
            sync_number += 1
        else:
            count_row["同步2"] = "FF"

    for stock_rows in stock_rows_by_simple_number.values():
        for stock_row in stock_rows:
            if not clean_text(stock_row.get("同步1", "")):
                stock_row["同步1"] = "SS"


def append_df5_group_summary_rows(df5: pd.DataFrame) -> pd.DataFrame:
    rows = df5.to_dict("records")
    columns = list(df5.columns)
    output_rows: list[dict[str, object]] = []
    group_rows_by_material_key: dict[str, list[dict[str, object]]] = {}

    for row in rows:
        material_key = get_df5_row_material_key(row)
        group_rows_by_material_key.setdefault(material_key, []).append(row)

    for group_rows in group_rows_by_material_key.values():
        apply_quantity_consistency_to_group(group_rows)
        sorted_group_rows = sort_df5_group_rows(group_rows)
        apply_sync_numbers_to_group(sorted_group_rows)
        output_rows.extend(sorted_group_rows)
        output_rows.append(make_df5_summary_row(columns, sorted_group_rows))

    return pd.DataFrame(output_rows, columns=columns)


def get_group_status(statuses: pd.Series) -> str:
    cleaned_statuses = set(statuses.map(clean_text))
    if "批次号不匹配" in cleaned_statuses:
        return "批次号不匹配"
    if "账实一致" in cleaned_statuses:
        return "账实一致"
    if "仅实物" in cleaned_statuses:
        return "仅实物"
    return "仅账面"


def summarize_df5_results(df5: pd.DataFrame) -> dict[str, int]:
    material_keys = df5.apply(get_df5_row_material_key, axis=1)
    summary = {
        "账实一致": 0,
        "批次号不匹配": 0,
        "仅实物": 0,
        "仅账面": 0,
    }

    grouped_statuses = df5.loc[material_keys != ""].groupby(material_keys[material_keys != ""])["状态"]
    for _, statuses in grouped_statuses:
        summary[get_group_status(statuses)] += 1

    return summary


def log_df5_summary(df5: pd.DataFrame) -> None:
    summary = summarize_df5_results(df5)
    log("处理结果统计：")
    for status, count in summary.items():
        log(f"{status}：{count}个物料")


def build_df5(df2: pd.DataFrame, df4_dataframes: dict[str, pd.DataFrame]) -> pd.DataFrame:
    df5 = df2.copy()
    convert_column_to_number(df5, "数量(生产单位)")
    df5.insert(0, "数据类型", "账面")
    batch_insert_at = df5.columns.get_loc("批次") + 1
    df5.insert(batch_insert_at, "简易制造号1", df5["批次"].map(build_simple_manufacture_number))
    df5.insert(batch_insert_at + 1, "同步1", "")

    stock_columns = list(df5.columns)
    new_columns = [
        "图号",
        "实物物料编码",
        "实物数量",
        "实物制造号",
        "简易制造号2",
        "同步2",
        "数量一致性",
        "状态",
        "S4物料编码",
    ]
    for column in new_columns:
        if column not in df5.columns:
            df5[column] = ""
    df5 = df5[[*stock_columns, *new_columns]]

    rows = df5.to_dict("records")
    columns = list(df5.columns)
    exact_indexes_by_key: dict[tuple[str, str], list[int]] = {}
    material_indexes_by_code: dict[str, list[int]] = {}

    for index, row in enumerate(rows):
        material_code = clean_text(row.get("物料", ""))
        simple_number = clean_text(row.get("简易制造号1", ""))
        if not material_code:
            continue
        exact_indexes_by_key.setdefault((material_code, simple_number), []).append(index)
        material_indexes_by_code.setdefault(material_code, []).append(index)

    insert_rows_after_index: dict[int, list[dict[str, object]]] = {}
    append_rows: list[dict[str, object]] = []

    for df4_name, df4 in df4_dataframes.items():
        validate_columns(df4, ["图号", "物料编码", "实物数量", "实物制造号"], df4_name)
        for _, physical_row in df4.iterrows():
            drawing_number = clean_text(physical_row["图号"])
            material_code = clean_text(physical_row["物料编码"])
            if not material_code:
                continue

            quantity = physical_row["实物数量"]
            manufacture_number = clean_text(physical_row["实物制造号"])
            simple_manufacture_number = build_simple_manufacture_number(manufacture_number)

            exact_indexes = exact_indexes_by_key.get((material_code, simple_manufacture_number), [])
            if exact_indexes:
                row_index = find_first_empty_status_index(rows, exact_indexes)
                rows[row_index]["状态"] = "账实一致"
                new_row = make_blank_df5_row(columns)
                new_row["数据类型"] = "盘点"
                fill_physical_fields(
                    new_row,
                    drawing_number,
                    material_code,
                    quantity,
                    manufacture_number,
                    "账实一致",
                )
                insert_rows_after_index.setdefault(row_index, []).append(new_row)
                continue

            material_indexes = material_indexes_by_code.get(material_code, [])
            if material_indexes:
                new_row = make_blank_df5_row(columns)
                new_row["数据类型"] = "盘点"
                fill_physical_fields(
                    new_row,
                    drawing_number,
                    material_code,
                    quantity,
                    manufacture_number,
                    "批次号不匹配",
                )
                insert_rows_after_index.setdefault(material_indexes[-1], []).append(new_row)
                continue

            new_row = make_blank_df5_row(columns)
            new_row["数据类型"] = "盘点"
            fill_physical_fields(
                new_row,
                drawing_number,
                material_code,
                quantity,
                manufacture_number,
                "仅实物",
            )
            append_rows.append(new_row)

    output_rows: list[dict[str, object]] = []
    for index, row in enumerate(rows):
        output_rows.append(row)
        output_rows.extend(insert_rows_after_index.get(index, []))
    output_rows.extend(append_rows)

    df5_result = apply_status_to_material_groups(pd.DataFrame(output_rows, columns=columns))
    return append_df5_group_summary_rows(df5_result)


def get_df5_material_key(row, header_indexes: dict[str, int]) -> str:
    stock_material = clean_text(row[header_indexes["物料"]].value)
    physical_material = clean_text(row[header_indexes["实物物料编码"]].value)
    return stock_material or physical_material


def auto_adjust_column_width(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            text = clean_text(cell.value)
            if text:
                max_length = max(max_length, len(text))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def get_sync_cell_fill(value: object, only_marker: str, only_color: str) -> PatternFill | None:
    text = clean_text(value)
    if not text:
        return None
    if text == only_marker:
        return PatternFill(fill_type="solid", fgColor=only_color)
    if text.isdigit():
        color = DF5_SYNC_MATCH_COLORS[(int(text) - 1) % len(DF5_SYNC_MATCH_COLORS)]
        return PatternFill(fill_type="solid", fgColor=color)
    return None


def format_df5_sheet(worksheet) -> None:
    left_alignment = Alignment(horizontal="left", vertical="center")
    stock_count_border = Border(top=Side(style="dashed", color="000000"))
    summary_border = Border(bottom=Side(style="medium", color="7F9DB9"))
    headers = [clean_text(cell.value) for cell in worksheet[1]]
    header_indexes = {header: index for index, header in enumerate(headers)}
    validate_columns(
        pd.DataFrame(columns=headers),
        ["数据类型", "物料", "实物物料编码", "同步1", "同步2"],
        worksheet.title,
    )

    color_index = -1
    previous_material_key = ""
    previous_data_type = ""
    current_fill = PatternFill(fill_type="solid", fgColor=DF5_GROUP_COLORS[0])
    current_summary_fill = PatternFill(fill_type="solid", fgColor=DF5_SUMMARY_COLORS[0])
    blank_fill = PatternFill(fill_type=None)

    for row_number, row in enumerate(worksheet.iter_rows(), start=1):
        add_summary_border = False
        add_stock_count_border = False
        data_type = ""
        if row_number == 1:
            fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
        else:
            data_type = clean_text(row[header_indexes["数据类型"]].value)
            material_key = get_df5_material_key(row, header_indexes)
            sync_cell_fills = {
                header_indexes["同步1"]: get_sync_cell_fill(
                    row[header_indexes["同步1"]].value,
                    "SS",
                    DF5_SYNC_STOCK_ONLY_COLOR,
                ),
                header_indexes["同步2"]: get_sync_cell_fill(
                    row[header_indexes["同步2"]].value,
                    "FF",
                    DF5_SYNC_PHYSICAL_ONLY_COLOR,
                ),
            }
            if data_type == "汇总":
                fill = current_summary_fill
                add_summary_border = row_number > 1
            elif not material_key:
                fill = blank_fill
            else:
                if material_key != previous_material_key:
                    color_index = (color_index + 1) % len(DF5_GROUP_COLORS)
                    current_fill = PatternFill(fill_type="solid", fgColor=DF5_GROUP_COLORS[color_index])
                    current_summary_fill = PatternFill(
                        fill_type="solid",
                        fgColor=DF5_SUMMARY_COLORS[color_index],
                    )

                previous_material_key = material_key
                fill = current_fill

            if data_type in {"账面", "盘点"}:
                add_stock_count_border = (
                    previous_data_type in {"账面", "盘点"}
                    and previous_data_type != data_type
                )
                previous_data_type = data_type
            elif data_type == "汇总":
                previous_data_type = ""

        for cell in row:
            cell.alignment = left_alignment
            cell_fill = sync_cell_fills.get(cell.column - 1) if row_number != 1 else None
            cell.fill = cell_fill or fill
            if add_summary_border:
                cell.border = summary_border
            elif add_stock_count_border:
                cell.border = stock_count_border

    auto_adjust_column_width(worksheet)


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "Sheet"
    base = cleaned[:31]
    sheet_name = base
    counter = 1
    while sheet_name in used_names:
        suffix = f"_{counter}"
        sheet_name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(sheet_name)
    return sheet_name


def build_output_path() -> Path:
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d-%H%M%S")
    return BASE_DIR / f"M5-3005库位盘点结果-{timestamp}.xlsx"


def main() -> None:
    configure_console_encoding()

    stock_file = find_latest_stock_file()
    log(f"读取库存文件：{stock_file}")
    df1 = pd.read_excel(stock_file, sheet_name=STOCK_SHEET_NAME)
    df2 = build_df2(df1)

    df3_dataframes: dict[str, pd.DataFrame] = {}
    df4_dataframes: dict[str, pd.DataFrame] = {}
    for count_file in list_count_files():
        code = extract_count_code(count_file)
        df3_name = f"df3-{code}"
        df4_name = f"df4-{code}"
        log(f"读取盘点文件：{count_file} -> {df3_name}, {df4_name}")
        raw_df3 = pd.read_excel(count_file, sheet_name=COUNT_SHEET_NAME)
        df3 = process_df3(raw_df3, df3_name)
        df3_dataframes[df3_name] = df3
        df4_dataframes[df4_name] = build_df4(df3, df4_name)

    df4_summary = build_df4_summary(df4_dataframes)

    log("生成 df5 对账结果")
    df5 = build_df5(df2, df4_dataframes)
    log_df5_summary(df5)

    output_path = build_output_path()
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name=safe_sheet_name("df1-Data", used_sheet_names), index=False)
        df2.to_excel(writer, sheet_name=safe_sheet_name("df2-库存抽取", used_sheet_names), index=False)
        for table_name, df3 in df3_dataframes.items():
            df3.to_excel(
                writer,
                sheet_name=safe_sheet_name(table_name, used_sheet_names),
                index=False,
            )
        df4_summary.to_excel(
            writer,
            sheet_name=safe_sheet_name("df4-汇总", used_sheet_names),
            index=False,
        )
        df5_sheet_name = safe_sheet_name("df5-对账结果", used_sheet_names)
        df5.to_excel(writer, sheet_name=df5_sheet_name, index=False)
        format_df5_sheet(writer.sheets[df5_sheet_name])

    log(f"处理完成，已生成：{output_path}")


if __name__ == "__main__":
    main()
