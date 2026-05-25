# -*- coding: utf-8 -*-
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable=None, *args, **kwargs):
        return iterable


INPUT_591E = Path("1-591E20260414.xlsx")
INPUT_MATERIAL_LIST = Path("2-物料列表.xlsx")
INPUT_NITRIDING_DETAIL = Path("3-氮化明细.xlsx")
OUTPUT_FILE_PREFIX = "处理后模具数据"

PROCESS_SUFFIXES = {"MC", "MR", "MN", "MT"}

DF2_COLUMNS = [
    "成品物料号",
    "成品物料号-物料冻结状态",
    "成品物料号-物料BOM状态",
    "是否为新增物料",
    "粗加工",
    "粗加工-物料冻结状态",
    "粗加工-物料BOM状态",
    "热处理",
    "热处理-物料冻结状态",
    "热处理-物料BOM状态",
    "氮化",
    "氮化-物料冻结状态",
    "氮化-物料BOM状态",
]

RESULT_GROUP_FILLS = {
    ("成品物料号", "成品物料号-物料冻结状态", "成品物料号-物料BOM状态"): "DDEBF7",
    ("粗加工", "粗加工-物料冻结状态", "粗加工-物料BOM状态"): "E2F0D9",
    ("热处理", "热处理-物料冻结状态", "热处理-物料BOM状态"): "FFF2CC",
    ("氮化", "氮化-物料冻结状态", "氮化-物料BOM状态"): "FCE4D6",
}
PROCESS_ROW_FILL_COLOR = "E7E6E6"


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def log(message: str) -> None:
    print(message, flush=True)


def progress(iterable, desc: str, total: int | None = None, colour: str = "green"):
    return tqdm(
        iterable,
        desc=desc,
        total=total,
        unit="行",
        dynamic_ncols=True,
        file=sys.stdout,
        colour=colour,
    )


def require_columns(df: pd.DataFrame, columns: list[str], table_name: str) -> None:
    missing = [col for col in columns if col not in df.columns]
    if missing:
        raise KeyError(f"{table_name} 缺少必要列: {missing}")


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_status_09(value: object) -> bool:
    text = clean_text(value)
    if text in {"09", "9", "9.0"}:
        return True
    return False


def is_zero(value: object) -> bool:
    if pd.isna(value):
        return False
    try:
        return float(value) == 0
    except (TypeError, ValueError):
        return clean_text(value) == "0"


def split_material_code(value: object) -> tuple[str, str]:
    text = clean_text(value)
    suffix = text[-2:].upper() if len(text) >= 2 else ""
    if suffix in PROCESS_SUFFIXES:
        return text[:-2], suffix
    return text, ""


def build_s4_material(value: object) -> str:
    mold_no = clean_text(value)
    if not mold_no:
        return ""
    return f"S-{mold_no}"


def add_blank_columns(df: pd.DataFrame, columns: list[str]) -> None:
    for col in columns:
        df[col] = ""


def build_output_path() -> Path:
    return Path(f"{OUTPUT_FILE_PREFIX}_{datetime.now():%Y%m%d}.xlsx")


def create_row_index(df: pd.DataFrame, key_col: str, desc: str) -> dict[str, int]:
    index_by_key: dict[str, int] = {}
    for row_idx in progress(df.index, desc=desc, total=len(df), colour="blue"):
        key = clean_text(df.at[row_idx, key_col])
        if key and key not in index_by_key:
            index_by_key[key] = row_idx
    return index_by_key


def mark_df1_status(df1: pd.DataFrame) -> None:
    add_blank_columns(df1, ["是否冻结物料", "是否缺少BOM信息"])

    for row_idx in progress(df1.index, desc="标记冻结/BOM状态", total=len(df1), colour="cyan"):
        if is_status_09(df1.at[row_idx, "基本视图状态"]):
            df1.at[row_idx, "是否冻结物料"] = "F"
        if is_zero(df1.at[row_idx, "BOM基本数量"]):
            df1.at[row_idx, "是否缺少BOM信息"] = "F"


def split_df1_material_codes(df1: pd.DataFrame) -> None:
    add_blank_columns(df1, ["子件号", "工序"])

    AA = ""
    BB = ""
    for row_idx in progress(df1.index, desc="拆分df1物料编码", total=len(df1), colour="cyan"):
        AA = ""
        BB = ""
        AA, BB = split_material_code(df1.at[row_idx, "物料编码"])
        df1.at[row_idx, "子件号"] = AA
        df1.at[row_idx, "工序"] = BB


def append_new_materials(df2: pd.DataFrame, df1A: pd.DataFrame) -> pd.DataFrame:
    existing = set()
    for row_idx in progress(df2.index, desc="建立df2子件号索引", total=len(df2), colour="blue"):
        child_no = clean_text(df2.at[row_idx, "子件号"])
        if child_no:
            existing.add(child_no)

    rows_to_append: list[dict[str, str]] = []
    for row_idx in progress(df1A.index, desc="比对新增物料", total=len(df1A), colour="yellow"):
        s4_material = clean_text(df1A.at[row_idx, "S4物料号"])
        if s4_material and s4_material not in existing:
            new_row = {col: "" for col in df2.columns}
            new_row["子件号"] = s4_material
            new_row["是否为新增物料"] = "是"
            rows_to_append.append(new_row)
            existing.add(s4_material)

    if rows_to_append:
        log(f"发现新增物料 {len(rows_to_append)} 条，追加到处理结果。")
        df2 = pd.concat([df2, pd.DataFrame(rows_to_append)], ignore_index=True)
    else:
        log("未发现需要追加的新增物料。")
    return df2


def fill_process_columns(df1: pd.DataFrame, df2: pd.DataFrame) -> None:
    row_index_by_child = create_row_index(df2, "子件号", "建立处理结果子件号索引")

    CC = ""
    DD = ""
    for row_idx in progress(df1.index, desc="回填工序与状态", total=len(df1), colour="green"):
        CC = ""
        DD = ""
        CC, DD = split_material_code(df1.at[row_idx, "物料编码"])
        if not CC or CC not in row_index_by_child:
            continue

        t = row_index_by_child[CC]
        frozen_status = df1.at[row_idx, "是否冻结物料"]
        bom_status = df1.at[row_idx, "是否缺少BOM信息"]

        if DD == "MC":
            df2.at[t, "粗加工"] = DD
            df2.at[t, "粗加工-物料冻结状态"] = frozen_status
            df2.at[t, "粗加工-物料BOM状态"] = bom_status
        elif DD == "MR":
            df2.at[t, "热处理"] = DD
            df2.at[t, "热处理-物料冻结状态"] = frozen_status
            df2.at[t, "热处理-物料BOM状态"] = bom_status
        elif DD in {"MN", "MT"}:
            df2.at[t, "氮化"] = DD
            df2.at[t, "氮化-物料冻结状态"] = frozen_status
            df2.at[t, "氮化-物料BOM状态"] = bom_status
        elif DD == "":
            df2.at[t, "成品物料号"] = CC
            df2.at[t, "成品物料号-物料冻结状态"] = frozen_status
            df2.at[t, "成品物料号-物料BOM状态"] = bom_status


def add_s4_material_column(df: pd.DataFrame, table_name: str) -> None:
    require_columns(df, ["模具号"], table_name)
    df["S4物料号"] = ""
    for row_idx in progress(df.index, desc=f"生成{table_name} S4物料号", total=len(df), colour="magenta"):
        df.at[row_idx, "S4物料号"] = build_s4_material(df.at[row_idx, "模具号"])


def fill_nitriding_record(df2: pd.DataFrame, df3: pd.DataFrame) -> None:
    df2["是否存在氮化记录"] = ""

    nitriding_index_by_s4 = create_row_index(df3, "S4物料号", "建立氮化记录索引")
    FF = ""
    matched_count = 0
    for row_idx in progress(df2.index, desc="匹配氮化记录", total=len(df2), colour="yellow"):
        FF = clean_text(df2.at[row_idx, "子件号"])
        if FF and FF in nitriding_index_by_s4:
            h = nitriding_index_by_s4[FF]
            df2.at[row_idx, "是否存在氮化记录"] = df3.at[h, "S4物料号"]
            matched_count += 1
    log(f"氮化记录匹配完成，共匹配 {matched_count} 条。")


def write_output(df1: pd.DataFrame, df1A: pd.DataFrame, df2: pd.DataFrame, df3: pd.DataFrame) -> Path:
    actual_output = build_output_path()
    try:
        write_excel(actual_output, df1, df1A, df2, df3)
    except PermissionError:
        actual_output = actual_output.with_name(
            f"{actual_output.stem}_{datetime.now():%Y%m%d_%H%M%S}{actual_output.suffix}"
        )
        log(f"无法覆盖当日输出文件，文件可能正在打开，改为另存为: {actual_output}")
        write_excel(actual_output, df1, df1A, df2, df3)
    return actual_output


def beautify_result_sheet(ws) -> None:
    log("正在美化处理结果: 四组列仅设置表头颜色。")
    headers = {cell.value: cell.column for cell in ws[1]}

    for columns, fill_color in RESULT_GROUP_FILLS.items():
        fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
        for column_name in columns:
            col_idx = headers.get(column_name)
            if col_idx is None:
                continue
            ws.cell(row=1, column=col_idx).fill = fill

    log("正在美化处理结果: 添加工序非空整行染色条件格式。")
    if ws.max_row < 2:
        log("处理结果无数据行，跳过整行染色规则。")
        return

    required_cols = ["粗加工", "热处理", "氮化"]
    missing_cols = [column_name for column_name in required_cols if column_name not in headers]
    if missing_cols:
        log(f"处理结果缺少列 {missing_cols}，跳过整行染色规则。")
        return

    row_fill = PatternFill(fill_type="solid", start_color=PROCESS_ROW_FILL_COLOR, end_color=PROCESS_ROW_FILL_COLOR)
    rough_letter = get_column_letter(headers["粗加工"])
    heat_letter = get_column_letter(headers["热处理"])
    nitriding_letter = get_column_letter(headers["氮化"])
    data_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    formula = f'OR(${rough_letter}2<>"",${heat_letter}2<>"",${nitriding_letter}2<>"")'
    ws.conditional_formatting.add(data_range, FormulaRule(formula=[formula], fill=row_fill))
    log("处理结果整行染色规则已添加。")


def write_excel(
    output_path: Path,
    df1: pd.DataFrame,
    df1A: pd.DataFrame,
    df2: pd.DataFrame,
    df3: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        log("正在写入 sheet: Data")
        df1.to_excel(writer, sheet_name="Data", index=False)
        log("正在写入 sheet: 物料生产记录")
        df1A.to_excel(writer, sheet_name="物料生产记录", index=False)
        log("正在写入 sheet: 处理结果")
        df2.to_excel(writer, sheet_name="处理结果", index=False)
        log("正在写入 sheet: 氮化记录")
        df3.to_excel(writer, sheet_name="氮化记录", index=False)
        beautify_result_sheet(writer.sheets["处理结果"])


def main() -> None:
    configure_console_encoding()

    log("开始读取 Excel 文件。")
    df1 = pd.read_excel(INPUT_591E, sheet_name="Data")
    df1A = pd.read_excel(INPUT_MATERIAL_LIST, dtype={"模具号": str})
    df3 = pd.read_excel(INPUT_NITRIDING_DETAIL, dtype={"模具号": str})

    require_columns(df1, ["物料编码", "基本视图状态", "BOM基本数量"], INPUT_591E.name)
    require_columns(df1A, ["模具号"], INPUT_MATERIAL_LIST.name)
    require_columns(df3, ["模具号"], INPUT_NITRIDING_DETAIL.name)

    log("处理 df1: 标记冻结状态、BOM状态，并拆分物料编码。")
    mark_df1_status(df1)
    split_df1_material_codes(df1)

    log("生成 df2: 提取子件号去重并增加结果列。")
    df2 = df1[["子件号"]].drop_duplicates().reset_index(drop=True)
    add_blank_columns(df2, DF2_COLUMNS)

    log("处理物料列表: 生成 S4物料号并追加新增物料。")
    add_s4_material_column(df1A, INPUT_MATERIAL_LIST.name)
    df2 = append_new_materials(df2, df1A)

    log("回填工序、冻结状态和 BOM 状态。")
    fill_process_columns(df1, df2)

    log("处理氮化明细: 生成 S4物料号并匹配氮化记录。")
    add_s4_material_column(df3, INPUT_NITRIDING_DETAIL.name)
    fill_nitriding_record(df2, df3)

    log("写入输出文件。")
    actual_output = write_output(df1, df1A, df2, df3)

    log("处理完成。")
    log(f"输出文件: {actual_output.resolve()}")
    log(f"Data 行数: {len(df1)}")
    log(f"物料生产记录 行数: {len(df1A)}")
    log(f"处理结果 行数: {len(df2)}")
    log(f"氮化记录 行数: {len(df3)}")


if __name__ == "__main__":
    main()
