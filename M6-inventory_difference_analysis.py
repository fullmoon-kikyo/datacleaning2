# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DF1_FILE = BASE_DIR / "03B_模具库盘点结果-手工账-260622.xlsx"
DF1_SHEET_NAME = "df4-汇总"
DF2_FILE = BASE_DIR / "03B_模具库盘点结果-手工账-260630.xlsx"
DF2_SHEET_NAME = 0
OUTPUT_PREFIX = "M6-盘点差异分析"


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def clean_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def validate_columns(df: pd.DataFrame, required_columns: list[str], table_name: str) -> None:
    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        raise KeyError(f"{table_name} 缺少必要列：{missing_columns}")


def find_existing_column(df: pd.DataFrame, candidate_columns: list[str], table_name: str) -> str:
    for column in candidate_columns:
        if column in df.columns:
            return column
    raise KeyError(f"{table_name} 缺少必要列，候选列名：{candidate_columns}")


def extract_suffix_from_filename(path: Path) -> str:
    return path.stem.rsplit("-", maxsplit=1)[-1]


def insert_or_replace_column(
    df: pd.DataFrame,
    after_column: str,
    new_column: str,
    values: pd.Series,
) -> pd.DataFrame:
    result = df.copy()
    if new_column in result.columns:
        result = result.drop(columns=[new_column])

    insert_at = result.columns.get_loc(after_column) + 1
    result.insert(insert_at, new_column, values)
    return result


def build_suite_material_number(value: object) -> str:
    text = clean_text(value)
    if "/" not in text:
        return text
    return text.split("/", maxsplit=1)[0].strip()


def build_s4_material_code(value: object) -> str:
    drawing_number = clean_text(value)
    if not drawing_number:
        return ""

    drawing_number = re.sub(r"[NFX]+$", "", drawing_number, flags=re.IGNORECASE)
    if not drawing_number:
        return ""
    return f"S-{drawing_number}"


def normalize_df6_source(df: pd.DataFrame) -> pd.DataFrame:
    material_code_column = find_existing_column(df, ["S4物料编码", "物料编码"], "df6")
    quantity_column = find_existing_column(df, ["数量", "实物数量"], "df6")

    result = df.copy()
    if material_code_column != "S4物料编码":
        result = result.rename(columns={material_code_column: "S4物料编码"})
    if quantity_column != "数量":
        result = result.rename(columns={quantity_column: "数量"})

    suite_material_number = result["S4物料编码"].map(build_suite_material_number)
    result = insert_or_replace_column(result, "S4物料编码", "成套物料号", suite_material_number)
    return result


def summarize_quantity_by_column(df: pd.DataFrame, group_column: str, table_name: str) -> pd.DataFrame:
    validate_columns(df, [group_column, "数量"], table_name)

    source = df[[group_column, "数量"]].copy()
    source[group_column] = source[group_column].map(clean_text)
    source["数量"] = pd.to_numeric(source["数量"], errors="coerce").fillna(0)
    source = source[source[group_column] != ""]

    result = (
        source.groupby(group_column, as_index=False, sort=True)["数量"]
        .sum()
        .reset_index(drop=True)
    )
    if (result["数量"] % 1 == 0).all():
        result["数量"] = result["数量"].astype(int)
    return result


def build_comparison_df(
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    group_column: str,
    right_key_column: str,
    right_quantity_column: str,
    left_table_name: str,
    right_table_name: str,
) -> pd.DataFrame:
    validate_columns(left_df, [group_column, "数量"], left_table_name)
    validate_columns(right_df, [group_column, "数量"], right_table_name)

    result = left_df.copy()
    lookup = right_df.set_index(group_column)
    matched_material_codes = result[group_column].map(
        lambda value: clean_text(value) if clean_text(value) in lookup.index else ""
    )
    matched_quantities = result[group_column].map(lookup["数量"])

    result = insert_or_replace_column(result, "数量", right_key_column, matched_material_codes)
    result = insert_or_replace_column(result, right_key_column, right_quantity_column, matched_quantities)

    left_keys = set(result[group_column].map(clean_text))
    right_only = right_df[~right_df[group_column].map(clean_text).isin(left_keys)].copy()
    if not right_only.empty:
        append_rows = pd.DataFrame(
            {
                group_column: "",
                "数量": "",
                right_key_column: right_only[group_column].map(clean_text),
                right_quantity_column: right_only["数量"],
            }
        )
        result = pd.concat([result, append_rows], ignore_index=True)

    result["盘点差异"] = (
        pd.to_numeric(result[right_quantity_column], errors="coerce").fillna(0)
        - pd.to_numeric(result["数量"], errors="coerce").fillna(0)
    )
    if (result["盘点差异"] % 1 == 0).all():
        result["盘点差异"] = result["盘点差异"].astype(int)
    return result


def build_df8a(df6a: pd.DataFrame, df7a: pd.DataFrame) -> pd.DataFrame:
    result = build_comparison_df(
        df6a,
        df7a,
        group_column="S4物料编码",
        right_key_column="df7A-S4物料编码",
        right_quantity_column="df7A-数量",
        left_table_name="df6A",
        right_table_name="df7A",
    )
    return result.rename(columns={"S4物料编码": "df6A-S4物料编码", "数量": "df6A-数量"})


def add_suite_quantities_to_df8a(
    df8a: pd.DataFrame,
    df6b: pd.DataFrame,
    df7b: pd.DataFrame,
) -> pd.DataFrame:
    validate_columns(df8a, ["df6A-S4物料编码", "df7A-S4物料编码"], "df8A")
    validate_columns(df6b, ["成套物料号", "数量"], "df6B")
    validate_columns(df7b, ["成套物料号", "数量"], "df7B")

    result = df8a.copy()
    source_material_codes = result["df6A-S4物料编码"].map(clean_text)
    fallback_material_codes = result["df7A-S4物料编码"].map(clean_text)
    suite_material_numbers = source_material_codes.where(source_material_codes != "", fallback_material_codes)
    result["成套物料号"] = suite_material_numbers.map(build_suite_material_number)

    df6b_lookup = df6b.set_index("成套物料号")["数量"]
    df7b_lookup = df7b.set_index("成套物料号")["数量"]
    result["df6B-成套数量"] = result["成套物料号"].map(df6b_lookup)
    result["df7B-成套数量"] = result["成套物料号"].map(df7b_lookup)
    result["成套-盘点差异"] = (
        pd.to_numeric(result["df7B-成套数量"], errors="coerce").fillna(0)
        - pd.to_numeric(result["df6B-成套数量"], errors="coerce").fillna(0)
    )
    if (result["成套-盘点差异"] % 1 == 0).all():
        result["成套-盘点差异"] = result["成套-盘点差异"].astype(int)
    return result


def build_df8b(df6b: pd.DataFrame, df7b: pd.DataFrame) -> pd.DataFrame:
    return build_comparison_df(
        df6b,
        df7b,
        group_column="成套物料号",
        right_key_column="df7B-成套物料号",
        right_quantity_column="df7B-数量",
        left_table_name="df6B",
        right_table_name="df7B",
    )


def color_comparison_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    left_key_column: str,
    right_key_column: str,
    difference_columns: list[str],
) -> None:
    validate_columns(df, [left_key_column, right_key_column, *difference_columns], sheet_name)

    worksheet = writer.sheets[sheet_name]
    fills = {
        "right_only": PatternFill(fill_type="solid", fgColor="00FFFF"),
        "row_positive": PatternFill(fill_type="solid", fgColor="9CC2E5"),
        "positive": PatternFill(fill_type="solid", fgColor="00B0F0"),
        "row_negative": PatternFill(fill_type="solid", fgColor="FFFF00"),
        "negative": PatternFill(fill_type="solid", fgColor="DA9694"),
        "zero": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    }
    difference_column_indexes = {
        column_name: df.columns.get_loc(column_name) + 1 for column_name in difference_columns
    }

    for row_index, row in df.iterrows():
        excel_row_index = row_index + 2
        if clean_text(row.get(left_key_column, "")) == "" and clean_text(row.get(right_key_column, "")) != "":
            row_fill = fills["right_only"]
        else:
            primary_difference = row[difference_columns[0]]
            if pd.isna(primary_difference):
                row_fill = None
            elif primary_difference > 0:
                row_fill = fills["row_positive"]
            elif primary_difference < 0:
                row_fill = fills["row_negative"]
            else:
                row_fill = fills["zero"]

        if row_fill is not None:
            for column_index in range(1, len(df.columns) + 1):
                worksheet.cell(row=excel_row_index, column=column_index).fill = row_fill

        for column_name, column_index in difference_column_indexes.items():
            difference = row[column_name]
            if pd.isna(difference) or difference == 0:
                continue
            if difference > 0:
                worksheet.cell(row=excel_row_index, column=column_index).fill = fills["positive"]
            elif difference < 0:
                worksheet.cell(row=excel_row_index, column=column_index).fill = fills["negative"]


def autofit_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    worksheet = writer.sheets[sheet_name]
    for index, column_name in enumerate(df.columns, start=1):
        values = [column_name, *df[column_name].head(200).astype(str).tolist()]
        max_length = max(len(value) for value in values if value != "nan")
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 40)


def left_align_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    alignment = Alignment(horizontal="left", vertical="center")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment


def main() -> None:
    configure_console_encoding()

    df1 = pd.read_excel(DF1_FILE, sheet_name=DF1_SHEET_NAME)
    df2 = pd.read_excel(DF2_FILE, sheet_name=DF2_SHEET_NAME)

    df1 = normalize_df6_source(df1)
    validate_columns(df2, ["图号", "数量"], "df2")

    s4_material_code = df2["图号"].map(build_s4_material_code)
    df2 = insert_or_replace_column(df2, "图号", "S4物料编码", s4_material_code)

    df2_suite_material_number = df2["S4物料编码"].map(build_suite_material_number)
    df2 = insert_or_replace_column(df2, "S4物料编码", "成套物料号", df2_suite_material_number)

    df6a = summarize_quantity_by_column(df1, "S4物料编码", "df6")
    df6b = summarize_quantity_by_column(df1, "成套物料号", "df6")
    df7a = summarize_quantity_by_column(df2, "S4物料编码", "df7")
    df7b = summarize_quantity_by_column(df2, "成套物料号", "df7")
    df8a = build_df8a(df6a, df7a)
    df8a = add_suite_quantities_to_df8a(df8a, df6b, df7b)
    df8b = build_df8b(df6b, df7b)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_file = BASE_DIR / f"{OUTPUT_PREFIX}-{timestamp}.xlsx"
    df7_suffix = extract_suffix_from_filename(DF2_FILE)
    df7_sheet_name = f"df7-{df7_suffix}"
    df7a_sheet_name = f"df7A-{df7_suffix}"
    df7b_sheet_name = f"df7B-{df7_suffix}"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="df6", index=False)
        df6a.to_excel(writer, sheet_name="df6A", index=False)
        df6b.to_excel(writer, sheet_name="df6B", index=False)
        df2.to_excel(writer, sheet_name=df7_sheet_name, index=False)
        df7a.to_excel(writer, sheet_name=df7a_sheet_name, index=False)
        df7b.to_excel(writer, sheet_name=df7b_sheet_name, index=False)
        df8a.to_excel(writer, sheet_name="df8A", index=False)
        df8b.to_excel(writer, sheet_name="df8B", index=False)
        autofit_columns(writer, "df6", df1)
        autofit_columns(writer, "df6A", df6a)
        autofit_columns(writer, "df6B", df6b)
        autofit_columns(writer, df7_sheet_name, df2)
        autofit_columns(writer, df7a_sheet_name, df7a)
        autofit_columns(writer, df7b_sheet_name, df7b)
        autofit_columns(writer, "df8A", df8a)
        autofit_columns(writer, "df8B", df8b)
        left_align_sheet(writer, "df8A")
        color_comparison_sheet(
            writer,
            "df8A",
            df8a,
            "df6A-S4物料编码",
            "df7A-S4物料编码",
            ["盘点差异", "成套-盘点差异"],
        )
        color_comparison_sheet(
            writer,
            "df8B",
            df8b,
            "成套物料号",
            "df7B-成套物料号",
            ["盘点差异"],
        )

    print(f"处理完成，输出文件：{output_file.name}")


if __name__ == "__main__":
    main()
